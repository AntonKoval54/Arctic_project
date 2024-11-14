import openpyxl as op
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import re
pattern = '\d{4}-\d{1,2}-\d{1,2}'

def process_data(filename):
  data_pre = {}
  glub_list = []
  count = 0
  with open(filename, "r") as f:
    lines = f.readlines()
  lines = [i for i in lines if i.find('VARIABLES')<0 and i.find('\n')!=0]
  print(lines)
  for i in range(0,len(lines),3):
    print('eeff',lines[i])
    if lines[i].find("ZONE")!=-1:
      try:
        match = re.findall(pattern, lines[i])[0]
        print("match", match)
        data_pre[match]=[]
      except:
        break
    if lines[i]!="\n" and lines[i].find("VARIABLES")<0:
      if count==0:# темпераутра
        glub_list = [float(x) for x in lines[i+1].split() if x]
        count+=1
      tempareture_list = [float(x) for x in lines[i+2].split() if x]
      print(tempareture_list)
      data_pre[match] =tempareture_list
  return glub_list, data_pre

def _to_xslx(temperature, data_temp_glub):
  wb = op.Workbook()
  if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
  ws = wb.create_sheet("45-1")
  for i, number in enumerate(temperature):
    ws.cell(row=1, column=i+2).value = number
  for i, (key, temperature_list) in enumerate(data_temp_glub.items()):
    ws.cell(row=i+2, column=1).value = key
    for num, tempr in enumerate(temperature_list):
      ws.cell(row=i+2, column=num+2).value = tempr
    
  wb.save("numbers_year_45_1_kovalsky.xlsx")
  print("yes")
  
      
if __name__ == "__main__":
  #g,d = process_data(r"D:\OneDrive\Рабочий стол\sk_uran_5sem\try2.txt")
  g,d = process_data(r"D:\OneDrive\Рабочий стол\sk_uran_5sem\12112024_SKVAZHINI\point4Zstep.dat")
  _to_xslx(g,d)